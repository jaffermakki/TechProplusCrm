import csv
import io
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def invoices_to_csv(invoices) -> str:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Invoice #", "Date", "Customer", "Subtotal", "Tax", "Total", "Payment Method"])
    for inv in invoices:
        writer.writerow([
            inv.invoice_number,
            inv.created_at.strftime("%Y-%m-%d %H:%M"),
            inv.customer.name if inv.customer else "Walk-in",
            f"{inv.subtotal:.2f}",
            f"{inv.tax_total:.2f}",
            f"{inv.total:.2f}",
            inv.payment_method,
        ])
    return buf.getvalue()


def month_end_workbook(invoices, top_products) -> io.BytesIO:
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Invoices"
    ws1.append(["Invoice #", "Date", "Customer", "Subtotal", "GST", "PST", "HST", "Total", "Payment Method"])
    for inv in invoices:
        ws1.append([
            inv.invoice_number,
            inv.created_at.strftime("%Y-%m-%d %H:%M"),
            inv.customer.name if inv.customer else "Walk-in",
            float(inv.subtotal),
            float(inv.gst_amount),
            float(inv.pst_amount),
            float(inv.hst_amount),
            float(inv.total),
            inv.payment_method,
        ])
    for col_idx in range(1, 10):
        ws1.column_dimensions[get_column_letter(col_idx)].width = 18

    ws2 = wb.create_sheet("Top Products")
    ws2.append(["Product", "Qty Sold", "Revenue"])
    for row in top_products:
        ws2.append([row["description"], row["qty"], row["revenue"]])
    for col_idx in range(1, 4):
        ws2.column_dimensions[get_column_letter(col_idx)].width = 24

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out
